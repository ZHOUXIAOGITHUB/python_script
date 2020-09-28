#-*- encoding=utf8 -*-
#Author:
#Time:  2020/6/9 14:31
#File:  Publice_Lib.py
import csv #导入csv
import openpyxl
class PubliceLib:
    ''' 冒泡排序, list 为所需要传输的列表'''
    def bubble_sort(self, list):
        for i in range(len(list)-1):  #循环次数比列表总数少一次
            #参加上次循环的不参与本次循环
            for j in range(len(list)-i-1):
                #如果列表list第j个数大于第j+1个数，则j和j+1换一个位子
                if list[j] > list[j+1]:
                    list[j], list[j+1] = list[j+1], list[j]
        # return list #此处return列表，调用方法没有收到返回的列表
        print(list)

    '''乘法表格用for遍历完成'''
    def Mulyiplication_Table_For(self):
        for i in range(1, 10): #从1开始循环到9停止，取1，不取10，不写"1"，默认从0开始
            for j in range(1,i+1 ):
                print("%d * %d = %d\t" %(i, j, i*j), end= "") #\表示制表符，end= ""表示换行
            print("")  #打印一次换行

    '''乘法表格用while循环'''
    def Mulyiplication_Table_While(self):
        i = 1
        while i <= 9:
            j = 1
            while j <= i:
                sum = j * i
                print("%dx%d=%2d" % (j, i, sum), end=("\t"))
                j += 1
            print()
            i += 1
        #此方法为搬过来的有待理解

class ExcelHelper:

    def get_excel_data(self, excel_file, sheet_name):
        '''
        获取excel文件数据
        :param excel_file: excel数据文件名称
        :param sheet_name: sheet表的名称
        :return: 返回excel的全部数据，
        如[('admin','123456'),('xiao1','123456'),('xiao2','123456'),('xiao3','123456')]
        '''
        # 获取excel文件
        wb = openpyxl.load_workbook(excel_file)
        # 指定sheet表
        get_sheet = wb[sheet_name]
        # 把每行的数据保存到一个列表
        get_one_line_list = []
        # 把所有行的元组数据保存到一个列表
        get_all_line_list = []
        par = True
        for get_line_tuple in get_sheet:
            # 获取每行单元格
            # print(get_line_tuple)
            # 过滤标题
            if par:
                par = False
                continue
            for get_one_cell in get_line_tuple:
                if get_one_cell.value == None:
                    # 当读取到单元格的值是空时，值转换成:''的格式。
                    get_one_cell.value = ''
                get_one_line_list.append(str(get_one_cell.value))
            # 获取完一行数据后，把数据所在的列表转化成元组
            get_one_line_tuple = tuple(get_one_line_list)
            get_all_line_list.append(get_one_line_tuple)
            # 清空列表，以便于存放另外一行数据，避免不同行数据保存到同一个列表中
            get_one_line_list.clear()
        return get_all_line_list

    def write_excel_data(self, excel_file, sheet_name, row, column, result):
        '''
        写入数据到excel文件
        :param excel_file: excel数据文件
        :param sheet_name: 文件的sheet表
        :param row: 需要修改数据所在的行
        :param column: 需要修改数据所在的列
        :param result: 修改的内容
        :return: None
        '''
        # 打开excel文件
        wb = openpyxl.load_workbook(excel_file)
        # 指定sheet表
        get_sheet = wb[sheet_name]
        # 修改第row行第column列的值
        get_sheet.cell(row, column).value = result
        # 保存关闭文件
        wb.save(excel_file)

    def get_data_rows(self, excel_file, sheet_name):
        '''
        获取用例数据的行数
        :param excel_file: 数据文件名
        :param sheet_name: sheet表名
        :return:
        '''
        wb = openpyxl.load_workbook(excel_file)
        # 指定sheet表
        get_sheet = wb[sheet_name]
        rows = get_sheet.max_row
        return rows

class CsvHelper:

    def get_csv_data(self, csv_file):
        '''
        读取csv文件数据
        :param csv_file: csv数据文件名称
        :return: csv文件数据
        '''
        with open(csv_file, mode='r', encoding='utf8') as csv_file:
            csv_data = csv.reader(csv_file)
            return csv_data










